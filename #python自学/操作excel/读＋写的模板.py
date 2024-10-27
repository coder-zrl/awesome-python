import xlrd
import openpyxl

# 创建表格
workbook = openpyxl.Workbook()  # 打开一个将写的文件，可以是本地文件
worksheet = workbook.create_sheet(index=0)  # 在将写的文件创建sheet
worksheet.cell(1, 1).value = '标题'
worksheet.cell(1, 2).value = '发文时间'
worksheet.cell(1, 3).value = '发文主体'
worksheet.cell(1, 4).value = '新闻正文'
row = 2

book = xlrd.open_workbook('address')#打开表格
sheet1 = book.sheets()[0]
nrows = sheet1.nrows
for nrow in range( nrows):
    #开始写
    worksheet.cell(row, 1).value = sheet1.cell(nrow, 0).value
    worksheet.cell(row, 2).value = sheet1.cell(nrow, 1).value
    worksheet.cell(row, 3).value = sheet1.cell(nrow, 2).value
    worksheet.cell(row, 4).value = sheet1.cell(nrow, 3).value
    row += 1
workbook.save('name.xlsx')